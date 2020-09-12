import openpyxl
from openpyxl import load_workbook

#reading excel
loc = ("drill list.xlsx") 
wBook = load_workbook(loc) 
wSheet=wBook['Sheet1']
rows = wSheet.max_row
columns = wSheet.max_column

#read cell per cell
ForShowing = input("column number to be shown")
ForTesting = input("column number for testing")
show = []
testing = []
for rownum in range(rows):	
	show.append((wSheet.cell(row=rownum+1, column=int(ForShowing))).value)
	testing.append((wSheet.cell(row=rownum+1, column=int(ForTesting))).value)
	
